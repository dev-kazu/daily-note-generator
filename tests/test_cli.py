from __future__ import annotations

from openpyxl import Workbook, load_workbook

from daily_note.cli import main


def test_cli_writes_note_from_argument(tmp_path):
    workbook_path = tmp_path / "notes.xlsx"

    exit_code = main(["--file", str(workbook_path), "--note", "hello"])

    workbook = load_workbook(workbook_path)
    worksheet = workbook["Notes"]

    assert exit_code == 0
    assert worksheet.max_row == 2
    assert worksheet.cell(row=2, column=4).value == "hello"


def test_cli_prompts_for_note_when_argument_is_missing(tmp_path, monkeypatch):
    workbook_path = tmp_path / "notes.xlsx"
    monkeypatch.setattr("builtins.input", lambda prompt: "prompted note")

    exit_code = main(["--file", str(workbook_path)])

    workbook = load_workbook(workbook_path)
    worksheet = workbook["Notes"]

    assert exit_code == 0
    assert worksheet.cell(row=2, column=4).value == "prompted note"


def test_cli_dry_run_does_not_create_workbook(tmp_path, capsys):
    workbook_path = tmp_path / "notes.xlsx"

    exit_code = main(["--file", str(workbook_path), "--note", "hello", "--dry-run"])

    captured = capsys.readouterr()
    assert exit_code == 0
    assert "Dry run: would add note" in captured.out
    assert str(workbook_path) in captured.out
    assert not workbook_path.exists()


def test_cli_dry_run_does_not_modify_existing_workbook(tmp_path, capsys):
    workbook_path = tmp_path / "notes.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Notes"
    worksheet.append(("date", "time", "weekday", "note"))
    worksheet.append(("2026-05-26", "09:00:00", "Tuesday", "existing"))
    workbook.save(workbook_path)

    exit_code = main(["--file", str(workbook_path), "--note", "hello", "--dry-run"])

    captured = capsys.readouterr()
    workbook = load_workbook(workbook_path)
    worksheet = workbook["Notes"]
    rows = list(worksheet.iter_rows(values_only=True))

    assert exit_code == 0
    assert "Dry run: would add note" in captured.out
    assert rows == [
        ("date", "time", "weekday", "note"),
        ("2026-05-26", "09:00:00", "Tuesday", "existing"),
    ]


def test_cli_rejects_empty_note(tmp_path, capsys):
    workbook_path = tmp_path / "notes.xlsx"

    exit_code = main(["--file", str(workbook_path), "--note", " "])

    captured = capsys.readouterr()
    assert exit_code == 1
    assert "note must not be empty" in captured.err
    assert not workbook_path.exists()


def test_cli_dry_run_rejects_empty_note_without_creating_workbook(tmp_path, capsys):
    workbook_path = tmp_path / "notes.xlsx"

    exit_code = main(["--file", str(workbook_path), "--note", " ", "--dry-run"])

    captured = capsys.readouterr()
    assert exit_code == 1
    assert "note must not be empty" in captured.err
    assert not workbook_path.exists()


def test_cli_dry_run_rejects_unexpected_header(tmp_path, capsys):
    workbook_path = tmp_path / "notes.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Notes"
    worksheet.append(("wrong", "header"))
    workbook.save(workbook_path)

    exit_code = main(["--file", str(workbook_path), "--note", "hello", "--dry-run"])

    captured = capsys.readouterr()
    workbook = load_workbook(workbook_path)
    worksheet = workbook["Notes"]

    assert exit_code == 1
    assert "unexpected header" in captured.err
    assert worksheet.max_row == 1
    assert tuple(cell.value for cell in worksheet[1]) == ("wrong", "header")


def test_cli_rejects_unknown_timezone(tmp_path, capsys):
    workbook_path = tmp_path / "notes.xlsx"

    exit_code = main([
        "--file",
        str(workbook_path),
        "--timezone",
        "Invalid/Timezone",
        "--note",
        "hello",
    ])

    captured = capsys.readouterr()
    assert exit_code == 2
    assert "unknown timezone" in captured.err
    assert not workbook_path.exists()
