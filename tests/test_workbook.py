from __future__ import annotations

from datetime import datetime, timezone
from zoneinfo import ZoneInfo

import pytest
from openpyxl import Workbook, load_workbook

from daily_note.workbook import HEADER, append_note, build_entry


TOKYO = ZoneInfo("Asia/Tokyo")


def test_append_note_creates_workbook_with_header_and_entry(tmp_path):
    workbook_path = tmp_path / "daily_notes.xlsx"
    now = datetime(2026, 5, 26, 4, 5, 6, tzinfo=timezone.utc)

    entry = append_note(workbook_path, "first note", TOKYO, now)

    workbook = load_workbook(workbook_path)
    worksheet = workbook["Notes"]
    rows = list(worksheet.iter_rows(values_only=True))

    assert entry.date == "2026-05-26"
    assert entry.time == "13:05:06"
    assert entry.weekday == "Tuesday"
    assert rows == [
        HEADER,
        ("2026-05-26", "13:05:06", "Tuesday", "first note"),
    ]


def test_append_note_preserves_existing_rows(tmp_path):
    workbook_path = tmp_path / "daily_notes.xlsx"

    append_note(workbook_path, "first", TOKYO, datetime(2026, 5, 26, 0, 0, 0, tzinfo=TOKYO))
    append_note(workbook_path, "second", TOKYO, datetime(2026, 5, 26, 1, 2, 3, tzinfo=TOKYO))

    workbook = load_workbook(workbook_path)
    worksheet = workbook["Notes"]
    rows = list(worksheet.iter_rows(values_only=True))

    assert rows == [
        HEADER,
        ("2026-05-26", "00:00:00", "Tuesday", "first"),
        ("2026-05-26", "01:02:03", "Tuesday", "second"),
    ]


def test_append_note_creates_notes_sheet_when_missing(tmp_path):
    workbook_path = tmp_path / "daily_notes.xlsx"
    workbook = Workbook()
    workbook.active.title = "Other"
    workbook.save(workbook_path)

    append_note(workbook_path, "created sheet", TOKYO, datetime(2026, 5, 26, tzinfo=TOKYO))

    workbook = load_workbook(workbook_path)
    assert "Other" in workbook.sheetnames
    assert "Notes" in workbook.sheetnames
    assert workbook["Notes"].cell(row=2, column=4).value == "created sheet"


def test_append_note_rejects_unexpected_header(tmp_path):
    workbook_path = tmp_path / "daily_notes.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Notes"
    worksheet.append(("wrong", "header"))
    workbook.save(workbook_path)

    with pytest.raises(ValueError, match="unexpected header"):
        append_note(workbook_path, "note", TOKYO, datetime(2026, 5, 26, tzinfo=TOKYO))


def test_build_entry_rejects_empty_note():
    with pytest.raises(ValueError, match="note must not be empty"):
        build_entry("   ", TOKYO, datetime(2026, 5, 26, tzinfo=TOKYO))


def test_build_entry_rejects_multiline_note():
    with pytest.raises(ValueError, match="note must be one line"):
        build_entry("line one\nline two", TOKYO, datetime(2026, 5, 26, tzinfo=TOKYO))
