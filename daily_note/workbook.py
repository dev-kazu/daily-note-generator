from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from zipfile import BadZipFile
from zoneinfo import ZoneInfo

from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException


SHEET_NAME = "Notes"
HEADER = ("date", "time", "weekday", "note")
WEEKDAYS = ("Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday")


@dataclass(frozen=True)
class NoteEntry:
    date: str
    time: str
    weekday: str
    note: str


def build_entry(note: str, timezone: ZoneInfo, now: datetime | None = None) -> NoteEntry:
    _validate_note(note)

    current = _current_time(timezone, now)
    return NoteEntry(
        date=current.strftime("%Y-%m-%d"),
        time=current.strftime("%H:%M:%S"),
        weekday=WEEKDAYS[current.weekday()],
        note=note,
    )


def append_note(
    workbook_path: Path,
    note: str,
    timezone: ZoneInfo,
    now: datetime | None = None,
) -> NoteEntry:
    path = Path(workbook_path)
    entry = build_entry(note, timezone, now)
    workbook = _load_or_create_workbook(path)
    worksheet = _get_or_create_notes_sheet(workbook)

    _ensure_header(worksheet)
    worksheet.append((entry.date, entry.time, entry.weekday, entry.note))

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    return entry


def _validate_note(note: str) -> None:
    if not note.strip():
        raise ValueError("note must not be empty")
    if "\n" in note or "\r" in note:
        raise ValueError("note must be one line")


def _current_time(timezone: ZoneInfo, now: datetime | None) -> datetime:
    if now is None:
        return datetime.now(timezone)
    if now.tzinfo is None:
        return now.replace(tzinfo=timezone)
    return now.astimezone(timezone)


def _load_or_create_workbook(workbook_path: Path) -> Workbook:
    if not workbook_path.exists():
        workbook = Workbook()
        workbook.active.title = SHEET_NAME
        return workbook

    try:
        return load_workbook(workbook_path)
    except (BadZipFile, InvalidFileException) as error:
        raise ValueError(f"cannot read workbook: {workbook_path}") from error


def _get_or_create_notes_sheet(workbook: Workbook):
    if SHEET_NAME in workbook.sheetnames:
        return workbook[SHEET_NAME]
    return workbook.create_sheet(SHEET_NAME)


def _ensure_header(worksheet) -> None:
    existing = tuple(
        worksheet.cell(row=1, column=index).value for index in range(1, len(HEADER) + 1)
    )
    if existing == HEADER:
        return

    if worksheet.max_row == 1 and all(value is None for value in existing):
        for index, value in enumerate(HEADER, start=1):
            worksheet.cell(row=1, column=index, value=value)
        return

    raise ValueError(f"{SHEET_NAME} sheet has an unexpected header")
